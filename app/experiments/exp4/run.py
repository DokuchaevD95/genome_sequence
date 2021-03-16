import csv

from openpyxl import Workbook
from openpyxl.styles.fills import GradientFill, Stop, Color

from typing import List
from typing import NamedTuple
from utils import GenomesReader
from Bio.SeqRecord import SeqRecord
from experiments.exp3.run import Application as Exp3Mixin


class GenomeColor(NamedTuple):
    genome_id: str
    color: str


class Application(Exp3Mixin):
    ALPHABET = ['A', 'C', 'G', 'T']
    GENOMES_PATH = '../../genomes/'

    def __init__(self):
        reader = GenomesReader(self.GENOMES_PATH)
        self.genomes: List[SeqRecord] = list(reader)
        self.genome_colors = self.read_colors('corona_type.csv')

    def run(self):
        search_results = self.find_repeats(self.genomes)

        wb = Workbook()
        ws = wb.active

        self.add_headers(ws)

        # Сборка таблицы
        colors_table = [
            ['', *[gen.genome_id for gen in self.genome_colors]]
        ]
        for index, first_gen in enumerate(self.genome_colors):
            row = [first_gen.genome_id]
            for second_gen in self.genome_colors:
                first_color = Color(self.get_color(first_gen.genome_id), type='rgba')
                second_color = Color(self.get_color(second_gen.genome_id), type='rgba')
                row.append(GradientFill(
                    type='linear',
                    degree=45,
                    stop=(
                        Stop(first_color, 0),
                        Stop(second_color, 1)
                    )
                ))
            colors_table.append(row)

        # Запись таблицы
        for row_index, row in enumerate(colors_table):
            for col_index, value in enumerate(row):
                cell = ws.cell(row_index + 1, col_index + 1)
                if isinstance(value, GradientFill):
                    cell.fill = value
                else:
                    cell.value = value

        wb.save('results/result.xlsx')

    def add_headers(self, ws):
        for index, genome in enumerate(self.genomes):
            ws.cell(1, index + 2)
            ws.cell(index + 2, 1).value = genome.id

    def get_color(self, genome_id: str) -> str:
        result = None
        for gen_color in self.genome_colors:
            if gen_color.genome_id == genome_id:
                result = gen_color.color
        return result

    @staticmethod
    def read_colors(file_name: str) -> List[GenomeColor]:
        result = []
        with open(file_name) as file:
            csv_reader = csv.reader(file)
            for row in csv_reader:
                result.append(GenomeColor(
                    genome_id=row[0],
                    color=row[1]
                ))
        return result


if __name__ == '__main__':
    Application().run()